from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

"""It solves two problems occurring in normal openpyxl column width reading:
    1. When file have column dimensions written in ranges - normal openpyxl width reading returns correct
    width value only for first column in range for next ones in range it returns default width
    so there is no way of knowing if column really have default width or have custom width but is just written in range
    
    2. For some reason reading column width using regular openpyxl function (here using default_width_source = 2)
    it gives value of 13 for default width"""

def get_correct_custom_column_widths(sheet):
    """ Reading real custom column dimensions.
    It,s better than normal openpyxl column dimension reading when xlsx have dimensions written in ranges.
    It reads dimensions only for those columns that have it saved in the file,
    so columns without custom width saved (so default ones) will be skipped"""

    custom_column_widths = {}
    sheet_col_dimensions = sheet.column_dimensions
    for start_col_letter in sheet_col_dimensions:
        start_col_dimensions = sheet_col_dimensions[start_col_letter]
        for col_idx in range(start_col_dimensions.min, start_col_dimensions.max + 1):
            custom_column_widths[get_column_letter(col_idx)] = start_col_dimensions.width

    return custom_column_widths

def get_correct_column_widths(sheet, default_width_source = 1):
    """ Reading real, more accurate column widths
    default_width_source = 1 > (recommended) default width values will be obtained using sheet.sheet_format.baseColWidth
    default_width_source = 2 > default width values will be obtained using sheet.sheet_format.baseColWidth
    default_width_source != 1 and 2 > default width values will be None
    """
    custom_column_widths = get_correct_custom_column_widths(sheet)

    better_widths = {}
    for col in sheet.columns:
        col_letter = col[0].column_letter
        if col_letter in custom_column_widths:
            better_widths[col_letter] = custom_column_widths[col_letter]
        else:
            if default_width_source == 1:
                default_column_width = sheet.sheet_format.baseColWidth
            elif default_width_source == 2:
                default_column_width = sheet.column_dimensions[col_letter].width
            else:
                default_column_width = None
            better_widths[col_letter] = default_column_width

    return better_widths

# Example usage:
file_path = "dimensions_test.xlsx"  # Replace with your file path
wb = load_workbook(file_path)
sheet = wb.active
correct_widths = get_correct_column_widths(sheet)
print(correct_widths)
